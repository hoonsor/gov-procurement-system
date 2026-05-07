#!/usr/bin/env python3
"""
估價單產生器 (Excel)
自動生成政府採購估價單
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

def create_quotation_excel(tender_info, line_items, company_info):
    """產生估價單 Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "估價單"
    
    # 設定樣式
    header_font = Font(bold=True, size=14)
    title_font = Font(bold=True, size=11)
    center_align = Alignment(horizontal='center', vertical='center')
    right_align = Alignment(horizontal='right', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=11)
    
    # === 表頭 ===
    ws.merge_cells('A1:G1')
    ws['A1'] = "估  價  單"
    ws['A1'].font = Font(bold=True, size=18)
    ws['A1'].alignment = center_align
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"估價單號碼：ES-{datetime.now().strftime('%Y%m%d')}-001"
    ws['A2'].alignment = Alignment(horizontal='right')
    
    # === 公司資訊 ===
    ws['A4'] = "廠商名稱："
    ws['A4'].font = title_font
    ws['B4'] = company_info.get("name", "")
    
    ws['A5'] = "統一編號："
    ws['A5'].font = title_font
    ws['B5'] = company_info.get("tax_id", "")
    
    ws['A6'] = "聯絡電話："
    ws['A6'].font = title_font
    ws['B6'] = company_info.get("phone", "")
    
    ws['D4'] = "估價日期："
    ws['D4'].font = title_font
    ws['E4'] = datetime.now().strftime("%Y年%m月%d日")
    
    ws['D5'] = "有效期限："
    ws['D5'].font = title_font
    ws['E5'] = f"自估價日起 {company_info.get('validity_days', 30)} 天"
    
    ws['D6'] = "聯絡人："
    ws['D6'].font = title_font
    ws['E6'] = company_info.get("contact", "")
    
    # === 採購資訊 ===
    ws.merge_cells('A8:G8')
    ws['A8'] = "【採購案資訊】"
    ws['A8'].font = title_font
    
    ws['A9'] = "採購案號："
    ws['A9'].font = title_font
    ws['B9'] = tender_info.get("tender_id", "")
    
    ws['C9'] = "採購名稱："
    ws['C9'].font = title_font
    ws.merge_cells('D9:G9')
    ws['D9'] = tender_info.get("title", "")
    
    ws['A10'] = "預算金額："
    ws['A10'].font = title_font
    ws['B10'] = f"NT$ {tender_info.get('budget', 0):,}"
    
    ws['C10'] = "截止日期："
    ws['C10'].font = title_font
    ws['D10'] = tender_info.get("deadline", "")
    
    # === 報價明細 ===
    row_start = 12
    headers = ["項次", "項目名稱", "規格說明", "單位", "數量", "單價", "小計"]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row_start, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    # 資料列
    row = row_start + 1
    total = 0
    
    for idx, item in enumerate(line_items, 1):
        amount = item["quantity"] * item["unit_price"]
        total += amount
        
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=1).alignment = center_align
        
        ws.cell(row=row, column=2, value=item["name"]).border = thin_border
        ws.cell(row=row, column=3, value=item.get("spec", "")).border = thin_border
        ws.cell(row=row, column=4, value=item.get("unit", "式")).border = thin_border
        ws.cell(row=row, column=4).alignment = center_align
        
        ws.cell(row=row, column=5, value=item["quantity"]).border = thin_border
        ws.cell(row=row, column=5).alignment = center_align
        
        ws.cell(row=row, column=6, value=item["unit_price"]).border = thin_border
        ws.cell(row=row, column=6).number_format = '#,##0'
        ws.cell(row=row, column=6).alignment = right_align
        
        ws.cell(row=row, column=7, value=amount).border = thin_border
        ws.cell(row=row, column=7).number_format = '#,##0'
        ws.cell(row=row, column=7).alignment = right_align
        
        row += 1
    
    # 合計列
    row += 1
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="總計").font = Font(bold=True, size=12)
    ws.cell(row=row, column=1).alignment = center_align
    ws.cell(row=row, column=1).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.cell(row=row, column=1).border = thin_border
    
    ws.cell(row=row, column=7, value=total).font = Font(bold=True, size=12)
    ws.cell(row=row, column=7).number_format = '#,##0'
    ws.cell(row=row, column=7).alignment = right_align
    ws.cell(row=row, column=7).fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    ws.cell(row=row, column=7).border = thin_border
    
    # 備註
    row += 2
    ws.cell(row=row, column=1, value="備註：").font = title_font
    ws.merge_cells(f'B{row}:G{row}')
    ws.cell(row=row, column=2, value="1. 本估價單未註明者，依政府採購法相關規定辦理。")
    
    row += 1
    ws.merge_cells(f'B{row}:G{row}')
    ws.cell(row=row, column=2, value="2. 本估價包含運費、安裝及一年保固服務。")
    
    row += 1
    ws.merge_cells(f'B{row}:G{row}')
    ws.cell(row=row, column=2, value="3. 報價含稅金，發票開立二聯式或三聯式發票。")
    
    # 簽章欄
    row += 2
    ws.cell(row=row, column=1, value="廠商負責人：").font = title_font
    ws.cell(row=row, column=3, value="（簽章）")
    
    ws.cell(row=row, column=5, value="主管核章：").font = title_font
    ws.cell(row=row, column=7, value="（簽章）")
    
    # 調整欄寬
    widths = [8, 20, 35, 8, 8, 12, 12]
    for i, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    return wb

def save_quotation_excel(tender_info, line_items, company_info, filepath=None):
    """儲存估價單"""
    if filepath is None:
        filepath = Path(__file__).parent.parent / "templates" / f"估價單_{tender_info.get('tender_id', 'unknown')}.xlsx"
    
    filepath = Path(filepath)
    filepath.parent.mkdir(parents=True, exist_ok=True)
    
    wb = create_quotation_excel(tender_info, line_items, company_info)
    wb.save(str(filepath))
    
    print(f"✅ 估價單已產生: {filepath}")
    return filepath

# 預設報價項目
DEFAULT_LINE_ITEMS = [
    {"name": "核心交換器", "spec": "48埠Gigabit + 4埠10G光纖", "unit": "台", "quantity": 4, "unit_price": 85000},
    {"name": "路由器", "spec": "企業級防火牆路由器", "unit": "台", "quantity": 2, "unit_price": 120000},
    {"name": "網路線材", "spec": "Cat6A 遮蔽雙絞線", "unit": "箱", "quantity": 20, "unit_price": 3500},
    {"name": "光纖收發模組", "spec": "10G SFP+ 多模模組", "unit": "顆", "quantity": 16, "unit_price": 6500},
    {"name": "網路管理系統", "spec": "中央網管平台授權", "unit": "套", "quantity": 1, "unit_price": 180000},
    {"name": "安裝工程", "spec": "含布線、測試及教育訓練", "unit": "式", "quantity": 1, "unit_price": 250000},
    {"name": "保固服務", "spec": "三年保固及24小時支援", "unit": "年", "quantity": 3, "unit_price": 150000},
]

if __name__ == "__main__":
    from pathlib import Path
    
    templates_dir = Path(__file__).parent.parent / "templates"
    templates_dir.mkdir(parents=True, exist_ok=True)
    
    # 預設公司資料
    company = {
        "name": "冠宇資訊科技有限公司",
        "tax_id": "12345678",
        "phone": "02-2771-5566",
        "contact": "陳大明",
        "validity_days": 30
    }
    
    # 預設採購資料
    tender = {
        "tender_id": "TP-2026-001",
        "title": "校園網路設備更新採購案",
        "budget": 2500000,
        "deadline": "2026-05-25"
    }
    
    doc = save_quotation_excel(tender, DEFAULT_LINE_ITEMS, company)
    print(f"📊 估價單路徑: {doc}")